"""
Excel ledger workbook builder.

Generates a multi-sheet Excel workbook with:
- Entity Register (account-to-entity mapping)
- Parent Debt Ledger (check deposit tracking)
- General Ledger (all transactions, double-entry)
- Account Summary (statement counts, balances)
- P&L per business entity
- Personal Financial Statement

Uses financial modeling color conventions:
  Blue = inputs, Black = formulas, Green = cross-sheet, Yellow = needs attention
"""

import json
from collections import defaultdict
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

from ..config import Config, Entity


class LedgerBuilder:
    """Builds the financial ledger Excel workbook from config and extracted data."""

    def __init__(self, config: Config):
        self.config = config
        self.wb = Workbook()
        self._init_styles()

    def _init_styles(self):
        fmt = self.config.excel_format
        self.input_font = Font(color=fmt.input_color, name="Arial", size=10)
        self.formula_font = Font(color=fmt.formula_color, name="Arial", size=10)
        self.crossref_font = Font(color=fmt.crossref_color, name="Arial", size=10)
        self.bold_font = Font(bold=True, name="Arial", size=10)
        self.header_font = Font(bold=True, name="Arial", size=11, color=fmt.header_fg)
        self.gray_font = Font(name="Arial", size=9, color="666666")
        self.header_fill = PatternFill("solid", fgColor=fmt.header_bg)
        self.attention_fill = PatternFill("solid", fgColor=fmt.attention_bg)
        self.subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        self.currency_fmt = fmt.currency_format

    def _load_transactions(self):
        """Load extracted transaction data from JSON."""
        txn_path = self.config.output_dir / "all_transactions.json"
        if txn_path.exists():
            with open(txn_path) as f:
                return json.load(f)
        return None

    def build(self) -> Workbook:
        """Build all sheets."""
        self.wb.remove(self.wb.active)
        self.txn_data = self._load_transactions()

        self._build_entity_register()
        self._build_parent_debt_ledger()
        self._build_general_ledger()
        self._build_account_summary()

        for entity in self.config.business_entities():
            self._build_pl(entity)
            self._build_balance_sheet(entity)

        self._build_personal_financial_stmt()

        return self.wb

    def save(self, path: Optional[str] = None):
        """Save workbook to file."""
        output = Path(path) if path else self.config.ledger_path
        output.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output))
        return str(output)

    def _build_entity_register(self):
        """Entity Register sheet -- maps accounts to legal entities."""
        ws = self.wb.create_sheet("Entity Register")

        ws.cell(row=1, column=1, value="ENTITY REGISTER").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"{self.config.case_name} | {self.config.case_number}").font = self.gray_font

        headers = ["Entity", "Type", "Account ID", "Institution", "Account Type", "Label", "Parser"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row = 5
        for entity in self.config.entities:
            for acct in entity.accounts:
                ws.cell(row=row, column=1, value=entity.name).font = self.formula_font
                ws.cell(row=row, column=2, value=entity.type).font = self.formula_font
                ws.cell(row=row, column=3, value=acct.id).font = self.input_font
                ws.cell(row=row, column=4, value=acct.institution).font = self.formula_font
                ws.cell(row=row, column=5, value=acct.type).font = self.formula_font
                ws.cell(row=row, column=6, value=acct.label).font = self.formula_font
                ws.cell(row=row, column=7, value=acct.parser).font = self.gray_font
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = self.border
                row += 1

    def _build_parent_debt_ledger(self):
        """Parent Debt Ledger -- tracks check deposits (possible family loans)."""
        ws = self.wb.create_sheet("Parent Debt Ledger")

        ws.cell(row=1, column=1, value="UNSECURED DEBT -- LOANS FROM FAMILY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"{self.config.case_name} | {self.config.case_number}").font = self.gray_font

        headers = ["Date", "Description", "From", "To Account", "Amount", "Running Total", "Source Document"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Pre-fill 40 transaction rows with running total formulas
        for r in range(6, 46):
            ws.cell(row=r, column=5).number_format = self.currency_fmt
            ws.cell(row=r, column=6).number_format = self.currency_fmt
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = self.border

            if r == 6:
                ws.cell(row=r, column=6, value='=E6').font = self.formula_font
            else:
                ws.cell(row=r, column=6, value=f'=IF(E{r}="","",F{r-1}+E{r})').font = self.formula_font

        # Total row
        ws.cell(row=47, column=1, value="TOTAL UNSECURED DEBT TO FAMILY").font = self.bold_font
        ws.cell(row=47, column=1).fill = self.subtotal_fill
        ws.cell(row=47, column=5, value='=SUM(E6:E46)').font = self.bold_font
        ws.cell(row=47, column=5).number_format = self.currency_fmt

    def _build_general_ledger(self):
        """General Ledger -- all transactions across all accounts, populated from extraction data."""
        ws = self.wb.create_sheet("General Ledger")

        ws.cell(row=1, column=1, value="GENERAL LEDGER").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"{self.config.case_name} | {self.config.case_number}").font = self.gray_font

        headers = ["Date", "Account", "Entity", "Type", "Description", "Debit", "Credit", "Category", "Balance", "Source"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        if not self.txn_data:
            ws.cell(row=5, column=1, value="No extraction data. Run: python scripts/extract_all.py").font = self.gray_font
            return

        # Collect all transactions, sorted by date
        all_txns = []
        for acct_id, acct in self.txn_data.get("accounts", {}).items():
            entity_name = acct.get("entity", "")
            for stmt in acct.get("statements", []):
                source = stmt.get("file", "")
                for dep in stmt.get("deposits", []):
                    all_txns.append({
                        "date": dep["date"], "account": acct_id, "entity": entity_name,
                        "type": "deposit", "desc": dep["description"],
                        "debit": 0, "credit": dep["amount"],
                        "category": dep.get("category", ""), "source": source,
                    })
                for wth in stmt.get("withdrawals", []):
                    all_txns.append({
                        "date": wth["date"], "account": acct_id, "entity": entity_name,
                        "type": "withdrawal", "desc": wth["description"],
                        "debit": wth["amount"], "credit": 0,
                        "category": wth.get("category", ""), "source": source,
                    })

        all_txns.sort(key=lambda t: t["date"])

        row = 5
        for txn in all_txns:
            ws.cell(row=row, column=1, value=txn["date"]).font = self.formula_font
            ws.cell(row=row, column=2, value=txn["account"]).font = self.input_font
            ws.cell(row=row, column=3, value=txn["entity"]).font = self.formula_font
            ws.cell(row=row, column=4, value=txn["type"]).font = self.formula_font
            ws.cell(row=row, column=5, value=txn["desc"]).font = self.formula_font
            if txn["debit"]:
                ws.cell(row=row, column=6, value=txn["debit"]).number_format = self.currency_fmt
            if txn["credit"]:
                ws.cell(row=row, column=7, value=txn["credit"]).number_format = self.currency_fmt
            ws.cell(row=row, column=8, value=txn["category"]).font = self.gray_font
            ws.cell(row=row, column=10, value=txn["source"]).font = self.gray_font
            for c in range(1, 11):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        # Totals
        ws.cell(row=row + 1, column=1, value="TOTALS").font = self.bold_font
        ws.cell(row=row + 1, column=1).fill = self.subtotal_fill
        ws.cell(row=row + 1, column=6, value=f'=SUM(F5:F{row-1})').font = self.bold_font
        ws.cell(row=row + 1, column=6).number_format = self.currency_fmt
        ws.cell(row=row + 1, column=7, value=f'=SUM(G5:G{row-1})').font = self.bold_font
        ws.cell(row=row + 1, column=7).number_format = self.currency_fmt

        ws.cell(row=2, column=3, value=f"{len(all_txns)} transactions").font = self.gray_font

    def _build_account_summary(self):
        """Account Summary -- one row per account with real extraction data."""
        ws = self.wb.create_sheet("Account Summary")

        ws.cell(row=1, column=1, value="ACCOUNT SUMMARY").font = Font(bold=True, size=14)

        headers = ["Account", "Entity", "Institution", "Type", "Label", "Status",
                    "Date Range", "Statements", "Deposits", "Withdrawals",
                    "Opening Bal", "Closing Bal", "Total In", "Total Out"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row = 4
        for entity in self.config.entities:
            for acct in entity.accounts:
                acct_data = None
                if self.txn_data:
                    acct_data = self.txn_data.get("accounts", {}).get(acct.id)

                ws.cell(row=row, column=1, value=acct.id).font = self.input_font
                ws.cell(row=row, column=2, value=entity.name).font = self.formula_font
                ws.cell(row=row, column=3, value=acct.institution).font = self.formula_font
                ws.cell(row=row, column=4, value=acct.type).font = self.formula_font
                ws.cell(row=row, column=5, value=acct.label).font = self.formula_font

                if acct_data:
                    stmts = acct_data.get("statements", [])
                    totals = acct_data.get("totals", {})
                    periods = [(s["period_start"], s["period_end"]) for s in stmts if s.get("period_start")]

                    if periods:
                        earliest = min(p[0] for p in periods)
                        latest = max(p[1] for p in periods)
                        ws.cell(row=row, column=7, value=f"{earliest} to {latest}").font = self.formula_font

                    # Check if account is closed
                    last_stmt = stmts[-1] if stmts else {}
                    is_closed = last_stmt.get("closing_balance", -1) == 0 and "CLOSED" in last_stmt.get("file", "").upper()
                    ws.cell(row=row, column=6, value="Closed" if is_closed else "Active").font = self.formula_font

                    ws.cell(row=row, column=8, value=totals.get("statements", 0)).font = self.formula_font
                    ws.cell(row=row, column=9, value=totals.get("deposits", 0)).font = self.formula_font
                    ws.cell(row=row, column=10, value=totals.get("withdrawals", 0)).font = self.formula_font

                    if stmts:
                        ws.cell(row=row, column=11, value=stmts[0].get("opening_balance", 0))
                        ws.cell(row=row, column=12, value=stmts[-1].get("closing_balance", 0))
                    ws.cell(row=row, column=13, value=totals.get("deposit_total", 0))
                    ws.cell(row=row, column=14, value=totals.get("withdrawal_total", 0))
                else:
                    ws.cell(row=row, column=6, value="No data").font = self.gray_font

                for c in [11, 12, 13, 14]:
                    ws.cell(row=row, column=c).number_format = self.currency_fmt
                for c in range(1, 15):
                    ws.cell(row=row, column=c).border = self.border
                row += 1

    def _build_pl(self, entity: Entity):
        """P&L sheet for a business entity."""
        sheet_name = f"{entity.name} P&L"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = self.wb.create_sheet(sheet_name)

        ws.cell(row=1, column=1, value=f"{entity.name} -- PROFIT & LOSS").font = Font(bold=True, size=14)

        # Revenue section
        ws.cell(row=4, column=1, value="REVENUE").font = self.bold_font
        ws.cell(row=5, column=1, value="Gross Receipts").font = self.formula_font
        ws.cell(row=6, column=1, value="Other Income").font = self.formula_font
        ws.cell(row=7, column=1, value="TOTAL REVENUE").font = self.bold_font
        ws.cell(row=7, column=1).fill = self.subtotal_fill

        # Expense section
        ws.cell(row=9, column=1, value="EXPENSES").font = self.bold_font
        ws.cell(row=10, column=1, value="Operating Expenses").font = self.formula_font
        ws.cell(row=11, column=1, value="Management Compensation").font = self.formula_font
        ws.cell(row=12, column=1, value="Tax Payments").font = self.formula_font
        ws.cell(row=13, column=1, value="Other Expenses").font = self.formula_font
        ws.cell(row=14, column=1, value="TOTAL EXPENSES").font = self.bold_font
        ws.cell(row=14, column=1).fill = self.subtotal_fill

        # Net income
        ws.cell(row=16, column=1, value="NET INCOME").font = Font(bold=True, size=12)

        # Account balances
        ws.cell(row=18, column=1, value="ACCOUNT BALANCES").font = self.bold_font
        row = 19
        for acct in entity.accounts:
            ws.cell(row=row, column=1, value=f"{acct.label} ({acct.id})").font = self.formula_font
            ws.cell(row=row, column=2).number_format = self.currency_fmt
            ws.cell(row=row, column=3).number_format = self.currency_fmt
            row += 1
        ws.cell(row=row, column=1, value="TOTAL BUSINESS CASH").font = self.bold_font
        ws.cell(row=row, column=1).fill = self.subtotal_fill

    def _build_personal_financial_stmt(self):
        """Personal Financial Statement -- assets and liabilities."""
        ws = self.wb.create_sheet("Personal Financial Stmt")

        ws.cell(row=1, column=1, value="PERSONAL FINANCIAL STATEMENT").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"{self.config.case_name} | {self.config.case_number}").font = self.gray_font

        # Assets
        ws.cell(row=4, column=1, value="ASSETS").font = Font(bold=True, size=12)
        asset_headers = ["Asset", "Description", "Current Value", "DOM Value", "Separate/Marital", "Notes"]
        for i, h in enumerate(asset_headers, 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Pre-fill asset rows
        row = 6
        for entity in self.config.personal_entities():
            for acct in entity.accounts:
                ws.cell(row=row, column=1, value=f"{acct.institution} {acct.id}").font = self.formula_font
                ws.cell(row=row, column=2, value=acct.label).font = self.formula_font
                ws.cell(row=row, column=3).number_format = self.currency_fmt
                ws.cell(row=row, column=3).fill = self.attention_fill
                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = self.border
                row += 1

        # Total assets
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL ASSETS").font = self.bold_font
        ws.cell(row=total_row, column=1).fill = self.subtotal_fill
        ws.cell(row=total_row, column=3, value=f'=SUM(C6:C{total_row-1})').font = self.bold_font
        ws.cell(row=total_row, column=3).number_format = self.currency_fmt

        # Liabilities
        liab_start = total_row + 2
        ws.cell(row=liab_start, column=1, value="LIABILITIES").font = Font(bold=True, size=12)
        liab_headers = ["Liability", "Creditor", "Balance Owed", "Monthly Payment", "Secured/Unsecured", "Notes"]
        for i, h in enumerate(liab_headers, 1):
            cell = ws.cell(row=liab_start + 1, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        liab_data_start = liab_start + 2
        liab_rows = [
            "Attorney Fees", "Credit Cards", "Parent/Family Loans",
            "Medical/Evaluator", "Supervised Visitation", "Legal Costs", "Other",
        ]
        for i, name in enumerate(liab_rows):
            r = liab_data_start + i
            ws.cell(row=r, column=1, value=name).font = self.formula_font
            ws.cell(row=r, column=3).number_format = self.currency_fmt
            ws.cell(row=r, column=3).fill = self.attention_fill
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = self.border

        # Parent loans cross-ref
        parent_row = liab_data_start + 2  # "Parent/Family Loans"
        ws.cell(row=parent_row, column=3, value="='Parent Debt Ledger'!E47").font = self.crossref_font
        ws.cell(row=parent_row, column=3).fill = PatternFill()

        # Total liabilities
        total_liab_row = liab_data_start + len(liab_rows)
        ws.cell(row=total_liab_row, column=1, value="TOTAL LIABILITIES").font = self.bold_font
        ws.cell(row=total_liab_row, column=1).fill = self.subtotal_fill
        ws.cell(row=total_liab_row, column=3,
                value=f'=SUM(C{liab_data_start}:C{total_liab_row-1})').font = self.bold_font
        ws.cell(row=total_liab_row, column=3).number_format = self.currency_fmt

        # Net worth
        nw_row = total_liab_row + 2
        ws.cell(row=nw_row, column=1, value="NET WORTH (Assets - Liabilities)").font = Font(bold=True, size=12)
        ws.cell(row=nw_row, column=3, value=f'=C{total_row}-C{total_liab_row}').font = Font(bold=True, size=12)
        ws.cell(row=nw_row, column=3).number_format = self.currency_fmt

    def _build_balance_sheet(self, entity: Entity):
        """Balance Sheet for a business entity — derived from extracted transaction data."""
        sheet_name = f"{entity.name} Bal Sheet"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = self.wb.create_sheet(sheet_name)

        ws.cell(row=1, column=1, value=f"{entity.name} — BALANCE SHEET").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"{self.config.case_name} | {self.config.case_number}").font = self.gray_font

        if not self.txn_data:
            ws.cell(row=4, column=1, value="No extraction data available.").font = self.gray_font
            return

        # Gather account data for this entity
        entity_accounts = {a.id: a for a in entity.accounts}
        account_balances = {}
        account_totals = {}

        for acct_id, acct_conf in entity_accounts.items():
            acct_data = self.txn_data.get("accounts", {}).get(acct_id)
            if not acct_data:
                continue
            stmts = acct_data.get("statements", [])
            totals = acct_data.get("totals", {})
            last_stmt = stmts[-1] if stmts else {}

            account_balances[acct_id] = {
                "label": acct_conf.label,
                "opening": stmts[0].get("opening_balance", 0) if stmts else 0,
                "closing": last_stmt.get("closing_balance", 0),
                "first_period": stmts[0].get("period_start", "") if stmts else "",
                "last_period": last_stmt.get("period_end", ""),
            }
            account_totals[acct_id] = {
                "total_deposits": totals.get("deposit_total", 0),
                "total_withdrawals": totals.get("withdrawal_total", 0),
                "deposit_count": totals.get("deposits", 0),
                "withdrawal_count": totals.get("withdrawals", 0),
            }

        # Determine reporting period
        all_periods = [ab["last_period"] for ab in account_balances.values() if ab["last_period"]]
        report_date = max(all_periods) if all_periods else "Current"
        ws.cell(row=3, column=1, value=f"As of: {report_date}").font = self.gray_font

        # === ASSETS ===
        row = 5
        ws.cell(row=row, column=1, value="ASSETS").font = Font(bold=True, size=12)
        row += 1

        # Current Assets - Cash
        ws.cell(row=row, column=1, value="Current Assets").font = self.bold_font
        row += 1

        headers = ["Account", "Description", "Balance"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        cash_start = row
        total_cash = 0
        for acct_id, bal in account_balances.items():
            ws.cell(row=row, column=1, value=acct_id).font = self.input_font
            ws.cell(row=row, column=2, value=bal["label"]).font = self.formula_font
            ws.cell(row=row, column=3, value=bal["closing"]).number_format = self.currency_fmt
            ws.cell(row=row, column=3).font = self.input_font
            total_cash += bal["closing"]
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        ws.cell(row=row, column=1, value="TOTAL CASH").font = self.bold_font
        ws.cell(row=row, column=1).fill = self.subtotal_fill
        ws.cell(row=row, column=3, value=f'=SUM(C{cash_start}:C{row-1})').font = self.bold_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        total_cash_row = row
        row += 2

        # Other assets (placeholder)
        ws.cell(row=row, column=1, value="Other Assets").font = self.bold_font
        row += 1
        ws.cell(row=row, column=1, value="Equipment / Property").font = self.formula_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        ws.cell(row=row, column=3).fill = self.attention_fill
        other_asset_row = row
        row += 1
        ws.cell(row=row, column=1, value="Accounts Receivable").font = self.formula_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        ws.cell(row=row, column=3).fill = self.attention_fill
        ar_row = row
        row += 2

        ws.cell(row=row, column=1, value="TOTAL ASSETS").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = self.subtotal_fill
        ws.cell(row=row, column=3, value=f'=C{total_cash_row}+C{other_asset_row}+C{ar_row}').font = Font(bold=True, size=12)
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        total_assets_row = row
        row += 2

        # === LIABILITIES ===
        ws.cell(row=row, column=1, value="LIABILITIES").font = Font(bold=True, size=12)
        row += 1

        ws.cell(row=row, column=1, value="Current Liabilities").font = self.bold_font
        row += 1
        liab_headers = ["Liability", "Description", "Amount"]
        for i, h in enumerate(liab_headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        liab_items = [
            "Accounts Payable",
            "Accrued Expenses",
            "Taxes Payable",
            "Other Current Liabilities",
        ]
        liab_start = row
        for item in liab_items:
            ws.cell(row=row, column=1, value=item).font = self.formula_font
            ws.cell(row=row, column=3).number_format = self.currency_fmt
            ws.cell(row=row, column=3).fill = self.attention_fill
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        ws.cell(row=row, column=1, value="TOTAL LIABILITIES").font = self.bold_font
        ws.cell(row=row, column=1).fill = self.subtotal_fill
        ws.cell(row=row, column=3, value=f'=SUM(C{liab_start}:C{row-1})').font = self.bold_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        total_liab_row = row
        row += 2

        # === EQUITY ===
        ws.cell(row=row, column=1, value="EQUITY").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Owner's Equity / Retained Earnings").font = self.formula_font
        ws.cell(row=row, column=3, value=f'=C{total_assets_row}-C{total_liab_row}').font = self.crossref_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        equity_row = row
        row += 2

        # === VERIFICATION ===
        ws.cell(row=row, column=1, value="VERIFICATION: Assets - Liabilities - Equity").font = self.gray_font
        ws.cell(row=row, column=3, value=f'=C{total_assets_row}-C{total_liab_row}-C{equity_row}').font = self.gray_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        row += 2

        # === CASH FLOW SUMMARY (from extracted data) ===
        ws.cell(row=row, column=1, value="CASH FLOW SUMMARY (from bank statements)").font = Font(bold=True, size=12)
        row += 1
        flow_headers = ["Account", "Total Deposits", "Total Withdrawals", "Net Flow", "Transactions"]
        for i, h in enumerate(flow_headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1

        flow_start = row
        for acct_id in entity_accounts:
            tots = account_totals.get(acct_id, {})
            dep_total = tots.get("total_deposits", 0)
            wth_total = tots.get("total_withdrawals", 0)
            dep_count = tots.get("deposit_count", 0)
            wth_count = tots.get("withdrawal_count", 0)

            ws.cell(row=row, column=1, value=acct_id).font = self.input_font
            ws.cell(row=row, column=2, value=dep_total).number_format = self.currency_fmt
            ws.cell(row=row, column=3, value=wth_total).number_format = self.currency_fmt
            ws.cell(row=row, column=4, value=dep_total - wth_total).number_format = self.currency_fmt
            ws.cell(row=row, column=4).font = self.crossref_font
            ws.cell(row=row, column=5, value=dep_count + wth_count).font = self.formula_font
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = self.border
            row += 1

        ws.cell(row=row, column=1, value="TOTAL").font = self.bold_font
        ws.cell(row=row, column=1).fill = self.subtotal_fill
        ws.cell(row=row, column=2, value=f'=SUM(B{flow_start}:B{row-1})').font = self.bold_font
        ws.cell(row=row, column=2).number_format = self.currency_fmt
        ws.cell(row=row, column=3, value=f'=SUM(C{flow_start}:C{row-1})').font = self.bold_font
        ws.cell(row=row, column=3).number_format = self.currency_fmt
        ws.cell(row=row, column=4, value=f'=B{row}-C{row}').font = self.bold_font
        ws.cell(row=row, column=4).number_format = self.currency_fmt
