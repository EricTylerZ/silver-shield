"""
Parent Debt Populator.

Identifies generic DEPOSIT entries (check deposits) across all personal accounts
and populates the Parent Debt Ledger sheet.

Key insight: payroll always shows as ACH/DIRECT DEPOSIT with employer name,
while parent loans (check deposits) show as just "DEPOSIT" with no description.
"""

from collections import defaultdict
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from ..config import Config
from ..categorizers.deposits import DepositCategorizer
from ..extractors.base import Statement, Transaction


class ParentDebtPopulator:
    """Populates Parent Debt Ledger from categorized deposits."""

    def __init__(self, config: Config):
        self.config = config
        self.categorizer = DepositCategorizer(config)

    def identify_parent_deposits(self, statements: dict[str, list[Statement]]) -> list[dict]:
        """
        Find all deposits flagged as possible parent debt across all accounts.

        Args:
            statements: Dict of account_id -> list of Statement objects

        Returns:
            List of deposit dicts sorted by date with account metadata
        """
        parent_deposits = []

        for account_id, stmts in statements.items():
            entity = self.config.get_entity_for_account(account_id)
            account = self.config.get_account(account_id)

            for stmt in stmts:
                for dep in stmt.deposits:
                    self.categorizer.categorize(dep)
                    if self.categorizer.is_possible_parent_debt(dep):
                        parent_deposits.append({
                            'date': dep.date,
                            'description': dep.description,
                            'amount': dep.amount,
                            'account': f"{account.institution} {account.id}" if account else account_id,
                            'entity': entity.name if entity else "Unknown",
                            'source': stmt.file_name,
                        })

        parent_deposits.sort(key=lambda x: x['date'])
        return parent_deposits

    def populate_ledger(self, ledger_path: str, parent_deposits: list[dict]) -> str:
        """
        Write parent deposits into the Parent Debt Ledger sheet.

        Returns summary string.
        """
        wb = load_workbook(ledger_path)
        ws = wb["Parent Debt Ledger"]

        blue = Font(color="0000FF", name="Arial", size=10)
        black = Font(color="000000", name="Arial", size=10)
        gray = Font(name="Arial", size=9, color="666666")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        fmt = self.config.excel_format.currency_format

        # Clear existing data rows (6-45)
        for row in range(6, 46):
            for c in range(1, 8):
                ws.cell(row=row, column=c).value = None

        # Populate
        for i, dep in enumerate(parent_deposits):
            row = 6 + i
            if row > 45:
                break
            ws.cell(row=row, column=1, value=dep['date']).font = blue
            ws.cell(row=row, column=2, value=dep['description']).font = black
            ws.cell(row=row, column=3, value="Family (presumed)").font = gray
            ws.cell(row=row, column=4, value=dep['account']).font = black
            ws.cell(row=row, column=5, value=dep['amount']).font = blue
            ws.cell(row=row, column=5).number_format = fmt
            ws.cell(row=row, column=6).number_format = fmt
            ws.cell(row=row, column=7, value=dep['source']).font = gray

            for c in range(1, 8):
                ws.cell(row=row, column=c).border = border

        wb.save(ledger_path)

        total = sum(d['amount'] for d in parent_deposits)
        return (
            f"Parent Debt Ledger populated: {len(parent_deposits)} transactions, "
            f"${total:,.2f} total, "
            f"{parent_deposits[0]['date']} to {parent_deposits[-1]['date']}"
            if parent_deposits else "No parent deposits identified"
        )
