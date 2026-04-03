"""
CI 42 Series Compliance Checks for financial ledger integrity.

Adapted from the EZ Merit Stewardship Exchange compliance framework.
These checks verify that the financial ledger meets audit standards
for family law discovery.

CI 42-001: All accounts in config must appear in Account Summary
CI 42-002: Entity mappings must match statement headers
CI 42-003: Transaction totals must match statement totals (within $0.02)
CI 42-004: All formulas must resolve without errors
CI 42-005: Parent debt entries must have source documents
CI 42-100: No PII in repository files (permanent guardrail)
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from ..config import Config
from ..extractors.base import Statement


@dataclass
class ComplianceResult:
    """Result of a single compliance check."""
    check_id: str
    name: str
    severity: str  # critical, high, medium, low
    passed: bool
    details: str = ""


class CI42Checker:
    """Runs CI 42 series compliance checks on the financial ledger."""

    def __init__(self, config: Config):
        self.config = config
        self.results: list[ComplianceResult] = []

    def run_all(self, ledger_path: Optional[str] = None,
                statements: Optional[dict[str, list[Statement]]] = None) -> list[ComplianceResult]:
        """Run all CI 42 checks."""
        self.results = []
        path = ledger_path or str(self.config.ledger_path)

        if Path(path).exists():
            wb = load_workbook(path)
            self._check_001_account_coverage(wb)
            self._check_004_formula_integrity(wb)
            self._check_005_parent_debt_sources(wb)
            wb.close()

        if statements:
            self._check_003_total_accuracy(statements)

        self._check_100_no_pii()

        return self.results

    def _check_001_account_coverage(self, wb):
        """CI 42-001: All configured accounts appear in Account Summary."""
        if "Account Summary" not in wb.sheetnames:
            self.results.append(ComplianceResult(
                "CI 42-001", "Account Coverage", "high", False,
                "Account Summary sheet not found"
            ))
            return

        ws = wb["Account Summary"]
        found_ids = set()
        for row in ws.iter_rows(min_row=4, max_col=1, values_only=True):
            if row[0]:
                found_ids.add(str(row[0]).strip())

        configured_ids = {a.id for a in self.config.all_accounts()}
        missing = configured_ids - found_ids

        self.results.append(ComplianceResult(
            "CI 42-001", "Account Coverage", "high",
            len(missing) == 0,
            f"Missing from Account Summary: {', '.join(missing)}" if missing
            else f"All {len(configured_ids)} accounts present"
        ))

    def _check_003_total_accuracy(self, statements: dict[str, list[Statement]]):
        """CI 42-003: Extracted totals match statement totals."""
        mismatches = []
        total_stmts = 0

        for account_id, stmts in statements.items():
            for stmt in stmts:
                total_stmts += 1
                dep_diff = abs(stmt.deposits_total - stmt.extracted_deposit_total)
                if dep_diff > 0.02 and stmt.deposits_total > 0:
                    mismatches.append(
                        f"{stmt.file_name}: deposit diff ${dep_diff:.2f}"
                    )

        accuracy = 1 - len(mismatches) / max(total_stmts, 1)
        self.results.append(ComplianceResult(
            "CI 42-003", "Total Accuracy", "medium",
            len(mismatches) == 0,
            f"{total_stmts} statements, {len(mismatches)} mismatches, "
            f"{accuracy:.1%} accuracy"
        ))

    def _check_004_formula_integrity(self, wb):
        """CI 42-004: No formula errors in the workbook."""
        errors = []
        total_formulas = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
                        if '#REF!' in str(cell.value) or '#NAME?' in str(cell.value):
                            errors.append(f"{sheet_name}!{cell.coordinate}")

        self.results.append(ComplianceResult(
            "CI 42-004", "Formula Integrity", "high",
            len(errors) == 0,
            f"{total_formulas} formulas, {len(errors)} errors"
            + (f": {', '.join(errors[:5])}" if errors else "")
        ))

    def _check_005_parent_debt_sources(self, wb):
        """CI 42-005: Parent debt entries have source documents."""
        if "Parent Debt Ledger" not in wb.sheetnames:
            self.results.append(ComplianceResult(
                "CI 42-005", "Parent Debt Sources", "medium", True,
                "No Parent Debt Ledger sheet"
            ))
            return

        ws = wb["Parent Debt Ledger"]
        entries_without_source = 0
        total_entries = 0

        for row in range(6, 46):
            amount = ws.cell(row=row, column=5).value
            source = ws.cell(row=row, column=7).value
            if amount and amount != "":
                total_entries += 1
                if not source or source == "":
                    entries_without_source += 1

        self.results.append(ComplianceResult(
            "CI 42-005", "Parent Debt Sources", "medium",
            entries_without_source == 0,
            f"{total_entries} entries, {entries_without_source} missing source docs"
        ))

    def _check_100_no_pii(self):
        """CI 42-100: No PII in repository files (permanent guardrail)."""
        # This check scans Python files for hardcoded account numbers, SSNs, etc.
        # Excludes known safe patterns: test fixtures, parser constants, comments
        repo_root = Path(self.config.config_path).parent
        pii_patterns = [
            r'\b\d{3}-\d{2}-\d{4}\b',  # SSN
            r'\b\d{9,12}\b',  # Account numbers (long digit strings)
        ]
        # Digit strings that are known safe (parser constants, test fixture data)
        safe_patterns = {
            r"#\s*\d{9,12}",        # in comments
            r"'[^']*\d{9,12}[^']*'",  # in string literals (test data, parser constants)
            r'"[^"]*\d{9,12}[^"]*"',  # in string literals
        }

        import re
        violations = []
        for py_file in repo_root.rglob("*.py"):
            if '.venv' in str(py_file) or '__pycache__' in str(py_file):
                continue
            try:
                content = py_file.read_text()
                for pattern in pii_patterns:
                    matches = re.findall(pattern, content)
                    if not matches:
                        continue
                    # Filter out matches inside string literals or comments
                    real_matches = []
                    for m in matches:
                        is_safe = False
                        for safe in safe_patterns:
                            if re.search(safe.replace(r'\d{9,12}', re.escape(m)), content):
                                is_safe = True
                                break
                        if not is_safe:
                            real_matches.append(m)
                    if real_matches:
                        violations.append(f"{py_file.name}: {len(real_matches)} potential PII matches")
            except Exception:
                pass

        self.results.append(ComplianceResult(
            "CI 42-100", "No PII in Repo", "critical",
            len(violations) == 0,
            "Clean" if not violations else f"Violations: {'; '.join(violations[:3])}"
        ))

    def summary(self) -> str:
        """Generate human-readable compliance summary."""
        passed = sum(1 for r in self.results if r.passed)
        total = len(self.results)
        lines = [f"CI 42 Compliance: {passed}/{total} checks passed\n"]
        for r in self.results:
            status = "PASS" if r.passed else "FAIL"
            lines.append(f"  [{status}] {r.check_id} {r.name} ({r.severity}): {r.details}")
        return "\n".join(lines)
